import re
import os
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import time

CBSE_CLASS_12_SUBJECT_CODES = {
    # Language Subjects
    "001": "ENGLISH ELECTIVE",
    "301": "ENGLISH CORE",
    "002": "HINDI ELECTIVE",
    "302": "HINDI CORE",
    "003": "URDU ELECTIVE",
    "303": "URDU CORE",
    "022": "SANSKRIT ELECTIVE",
    "322": "SANSKRIT CORE",
    "104": "PUNJABI",
    "105": "BENGALI",
    "106": "TAMIL",
    "107": "TELUGU",
    "108": "SINDHI",
    "109": "MARATHI",
    "110": "GUJARATI",
    "111": "MANIPURI",
    "112": "MALAYALAM",
    "113": "ODIA",
    "114": "ASSAMESE",
    "115": "KANNADA",
    "116": "ARABIC",
    "117": "TIBETAN",
    "118": "FRENCH",
    "120": "GERMAN",
    "121": "RUSSIAN",
    "123": "PERSIAN",
    "124": "NEPALI",
    "125": "LIMBOO",
    "126": "LEPCHA",
    "189": "TELUGU TELANGANA",
    "192": "BODO",
    "193": "TANGKHUL",
    "194": "JAPANESE",
    "195": "BHUTIA",
    "196": "SPANISH",
    "197": "KASHMIRI",
    "198": "MIZO",
    "199": "BAHASA MELAYU",

    # Academic Subjects
    "027": "HISTORY",
    "028": "POLITICAL SCIENCE",
    "029": "GEOGRAPHY",
    "030": "ECONOMICS",
    "031": "CARNATIC MUSIC (VOCAL)",
    "032": "CARNATIC MUSIC( MELODIC INSTRUMENTS).",
    "033": "CARNATIC MUSIC ( PERCUSSION INSTRUMENTS - MRIDANGAM)",
    "034": "HINDUSTANI MUSIC (VOCAL)",
    "035": "HINDUSTANI MUSIC ( MELODIC INSTRUMENTS).",
    "036": "HINDUSTANI MUSIC ( PERCUSSION INSTRUMENTS)",
    "037": "PSYCHOLOGY",
    "039": "SOCIOLOGY",
    "041": "MATHEMATICS",
    "042": "PHYSICS",
    "043": "CHEMISTRY",
    "044": "BIOLOGY",
    "045": "BIOTECHNOLOGY",
    "046": "ENGINEERING GRAPHICS",
    "048": "PHYSICAL EDUCATION",
    "049": "PAINTING",
    "050": "GRAPHICS",
    "051": "SCULPTURE",
    "052": "APPLIED/ COMMERCIAL ART",
    "054": "BUSINESS STUDIES",
    "055": "ACCOUNTANCY",
    "056": "KATHAK - DANCE",
    "057": "BHARATNATYAM - DANCE",
    "058": "KUCHIPUDI-DANCE",
    "059": "ODISSI - DANCE",
    "060": "MANIPURI - DANCE",
    "061": "KATHAKALI - DANCE",
    "064": "HOME SCIENCE",
    "265": "INFORMATICS PRACTICES (OLD)(Only for XII)",
    "065": "INFORMATICS PRACTICES (NEW)",
    "283": "COMPUTER SCIENCE (OLD) (Only for XII)",
    "083": "COMPUTER SCIENCE (NEW)",
    "066": "ENTREPRENEURSHIP",
    "073": "KNOWLEDGE TRADITION & PRACTICES OF INDIA",
    "074": "LEGAL STUDIES",
    "076": "NATIONAL CADET CORPS (NCC)",
}

CBSE_CLASS_10_SUBJECT_CODES = {
    # Language Subjects
    "002": "HINDI COURSE-A",
    "085": "HINDI COURSE-B",
    "184": "ENGLISH LANG & LIT.",
    "003": "URDU COURSE-A",
    "303": "URDU COURSE-B",
    "004": "PUNJABI",
    "005": "BENGALI",
    "006": "TAMIL",
    "007": "TELUGU",
    "008": "SINDHI",
    "009": "MARATHI",
    "010": "GUJARATI",
    "011": "MANIPURI",
    "012": "MALAYALAM",
    "013": "ODIA",
    "014": "ASSAMESE",
    "015": "KANNADA",
    "016": "ARABIC",
    "017": "TIBETAN",
    "018": "FRENCH",
    "020": "GERMAN",
    "021": "RUSSIAN",
    "023": "PERSIAN",
    "024": "NEPALI",
    "025": "LIMBOO",
    "026": "LEPCHA",
    "089": "TELUGU TELANGANA",
    "092": "BODO",
    "093": "TANGKHUL",
    "094": "JAPANESE",
    "095": "BHUTIA",
    "096": "SPANISH",
    "097": "KASHMIRI",
    "098": "MIZO",
    "099": "BAH ASA MELAYU",
    "122": "SANSKRIT",
    "131": "RAI",
    "132": "GURUNG",
    "133": "TAMANG",
    "134": "SHERPA",
    "136": "THAI",

    # Academic Subjects
    "041": "MATHEMATICS -STANDARD",
    "241": "MATHEMATICS -BASIC",
    "086": "SCIENCE",
    "087": "SOCIAL SCIENCE",

    # Additional Academic Subjects
    "031": "CARNATIC MUSIC (VOCAL)",
    "032": "CARNATIC MUSIC (MELODIC INSTRUMENTS)",
    "033": "CARNATIC MUSIC (PERCUSSION INSTRUMENTS)",
    "034": "HINDUSTANI MUSIC (VOCAL)",
    "035": "HINDUSTANI MUSIC (MELODIC INSTRUMENTS)",
    "036": "HINDUSTANI MUSIC (PERCUSSION INSTRUMENTS)",
    "049": "PAINTING",
    "064": "HOME SCIENCE",
    "076": "NATIONAL CADET CORPS (NCC)",
    "165": "COMPUTER APPLICATIONS",
    "154": "ELEMENTS OF BUSINESS",
    "254": "ELEMENTS OF BOOK KEEPING & ACCOUNTANCY",

    # Skill Subjects
    "401": "RETAILING",
    "402": "INFORMATION TECHNOLOGY",
    "403": "SECURITY",
    "404": "AUTOMOTIVE",
    "405": "INTRODUCTION TO FINANCIAL MARKETS",
    "406": "INTRODUCTION TO TOURISM",
    "407": "BEAUTY & WELLNESS",
    "408": "AGRICULTURE",
    "409": "FOOD PRODUCTION",
    "410": "FRONT OFFICE OPERATIONS",
    "411": "BANKING & INSURANCE",
    "412": "MARKETING & SALES",
    "413": "HEALTH CARE",
    "414": "APPAREL",
    "415": "MEDIA",
    "416": "MULTI SKILL FOUNDATION COURSE",
    "417": "ARTIFICIAL INTELLIGENCE",

}

# ======================
# REGEX VALUES
# ======================
# Change these values in case the format changes

subject_codes_regex = re.compile(r'(?<=\s)\d\d\d(?=\s)')
roll_no_regex = re.compile(r'\d{8,}')
gender_regex = re.compile(r'(?<=\s)[FM](?=\s)')
name_regex = re.compile(r'[A-Z]+')  # then remove "PASS"
marks_regex = re.compile(r'(?<=\s)\d\d\d(?=\s)')
pass_fail_regex = re.compile(r'PASS|COMP|ESSENTIAL REPEAT')
grades_regex = re.compile(r'[A-E][12]|(?<=\s)E(?=\s)')
AllSubjectNames = []


# ======================
# INPUT FILE FUNCTIONS
# ======================

def get_lines(file):
    """Takes the file path as parameter and returns a list of lines"""
    with open(file, 'r') as f:
        lines = f.readlines()
    return lines


def filter_lines(list_of_lines, mode):
    """Takes the list of lines and the format as input and removes the unnecessary lines
    Parameters:
        list_of_lines : list
        mode : str (either "10th" or "12th")
    Returns:
        filtered lines containing student data.
    """
    filtered_lines = []
    total_students = 0

    # Removing Empty Lines
    list_of_lines = [line for line in list_of_lines if line.strip() != '']

    for line_number, line in enumerate(list_of_lines):
        if "ABST" in line:
            continue
        if contains_student_data(line):
            total_students += 1
            filtered_lines.append(list_of_lines[line_number])
            filtered_lines.append(list_of_lines[line_number + 1])

    print("Found {} students.".format(total_students))
    return filtered_lines


def contains_student_data(line):
    """Takes a string as input and checks whether that string contains student's data
    Arguments:
        line : str
    Returns:
        True, if contains student's roll no and therefore the data
        False, if it doesn't contain student's roll no and therefore no data.
    """
    if roll_no_regex.search(line) is not None:
        return True
    return False


# ======================
# CONVERSION FUNCTIONS
# ======================

def convert_to_list_of_integer(list_of_strings):
    """Converts a list of strings to a list of integers"""
    return [int(i) for i in list_of_strings]


def convert_codes_to_subjects(list_of_codes, mode):
    """Converts a list of CBSE subject codes to a list of Corresponding Subjects"""
    if mode == '10th':
        return [CBSE_CLASS_10_SUBJECT_CODES[code] for code in list_of_codes]
    return [CBSE_CLASS_12_SUBJECT_CODES[code] for code in list_of_codes]


# ======================
# STRING DATA EXTRACTION FUNCTIONS
# ======================

def get_subject_codes(string_containing_subject_codes):
    """Searches the string for CBSE Subject Codes
    Parameters:
        string_containing_subject_codes : string that has subject codes
    Returns:
        List of all the subject codes
    """
    return subject_codes_regex.findall(string_containing_subject_codes)


def get_marks(string_containing_marks):
    """Searches the string for student's marks
    Parameters:
        string_containing_marks : str (containing marks)
    Returns:
        List of all the marks
    """
    marks_tuple = marks_regex.findall(string_containing_marks)
    return [int(x) for x in marks_tuple]


def get_name(string_containing_name):
    """Searches the string for student's marks
    Parameters:
        string_containing_name : str (containing the student's name)
    Returns:
        Name : str
    """

    return string_containing_name[13:64].strip()


def get_grades(string_containing_grades):
    """Searches the string for student's grades
    Parameters:
        string_containing_grades : str (containing grades)
    Returns:
        List of all the grades in the string
    """
    grades_tuple = grades_regex.findall(string_containing_grades)
    return list(grades_tuple)


def get_gender(string_containing_gender):
    """Searches the string for student's gender
    Parameters:
        string_containing_gender : A string that contains gender
    Returns:
       Gender : str
    """
    gender = gender_regex.search(string_containing_gender).group(0)
    return gender


def get_roll_no(string_containing_roll_no):
    """Searches the string for student's name
    Parameters:
        string_containing_roll_no : A string that contains student's Roll No
    Returns:
       roll no : int
    """
    roll_no = roll_no_regex.search(string_containing_roll_no).group(0)
    return roll_no


def get_result(string_containing_result):
    """Searches the string for student's result (PASS/FAIL/COMP/NECESSARY REPEAT)
    Parameters:
        string_containing_result : A string that contains student's Result
    Returns:
        result : str
    """
    result = pass_fail_regex.search(string_containing_result).group(0)
    return result


def get_marks_and_grades(first_line_without_compartment, second_line, subject_code_order):
    subject_codes_taken_by_student = get_subject_codes(first_line_without_compartment)
    grades_of_student = get_grades(second_line)
    marks_of_student = get_marks(second_line)
    marks_and_grades = [None] * 2 * len(subject_code_order)
    for index, subject_code in enumerate(subject_codes_taken_by_student):
        i = subject_code_order.index(subject_code)
        marks_and_grades[2 * i], marks_and_grades[2 * i + 1] = marks_of_student[index], grades_of_student[index]
    return marks_and_grades


def get_compartment_subjects(string, mode):
    if mode == "10th":
        compartment_subject_start = 127
    else:
        compartment_subject_start = 144

    compartment_subjects = string[compartment_subject_start:].strip().split()
    if len(compartment_subjects) > 0:
        compartment_subjects = [i + "(" + convert_codes_to_subjects([i], mode)[0] + ")" for i in compartment_subjects]
        return " ".join(compartment_subjects)
    else:
        return None


def get_unique_subject_codes(list_of_lines_containing_subject_codes):
    unique_subject_codes = set()
    for line in list_of_lines_containing_subject_codes:
        codes = get_subject_codes(line)
        for code in codes:
            unique_subject_codes.add(code)
    return list(unique_subject_codes)


def get_three_grades(line_containing_three_grades):
    return line_containing_three_grades[113:123].strip().split()


def get_absentees(input_file):
    file = open(input_file, "r")
    list_of_lines = file.readlines()
    absentees = []
    for line in list_of_lines:
        if "ABST" in line:
            absentees.append([get_roll_no(line), get_name(line)])
    return absentees

# ======================
# DATA VALIDATION FUNCTIONS
# ======================


def validate_input_path(file_name):
    """Checks whether the input file is valid or not
    Parameters:
        file_name : str
    Returns:
        True, if the file is valid
        False, if the file is invalid
    """
    if not os.path.isfile(file_name):
        print("File {} does not exist".format(file_name))
        return False
    if not (file_name.endswith(".txt") or file_name.endswith(".TXT")):
        print("File {} is not a text file".format(file_name))
        return False
    return True


def validate_output_path(file_name):
    """Checks whether the output file is valid or not
    Parameters:
        file_name : str
    Returns:
        True, if the file is valid
        False, if the file is invalid
    """
    if os.path.dirname(file_name) != "" and not os.path.exists(os.path.dirname(file_name)):
        print("ERROR: Directory {} does not exist. Please create the directory.".format(os.path.dirname(file_name)))
        return False
    return True


# ======================
# GENERAL FUNCTIONS
# ======================


def get_headers(unique_subject_codes, mode):
    headers_part_1 = ["Roll No", "Gender", "Name"]
    unique_subject_names = convert_codes_to_subjects(unique_subject_codes, mode)
    for subject in unique_subject_names:
        headers_part_1.extend([subject, "Grade"])
    if mode == "12th":
        headers_part_2 = ["GR1", "GR2", "GR3", "Result", "Best 5 percentage", "Compartment Subject"]
    else:
        headers_part_2 = ["Result", "Best 5 percentage", "Compartment Subject"]
    headers = headers_part_1 + headers_part_2
    return headers, unique_subject_names


def best_5_percent(marks_list):
    marks_list.sort(reverse=True)
    total_marks = 0
    for i in marks_list[:5]:
        total_marks += i
    return total_marks / 5


def append_title(ws_object, title, end_column, start_column=1, top_row=True):
    if top_row:
        ws_object.append([])
    ws_object.append([title])
    ws_object.cell(ws_object.max_row, column=start_column).font = Font(bold=True)
    ws_object.merge_cells(start_row=ws_object.max_row, start_column=start_column, end_row=ws_object.max_row,
                          end_column=end_column)


def save_wb(workbook, path):
    try:
        workbook.save(path)
    except PermissionError:
        print(
            "ERROR: Unable to save the excel file. Please change the output file name or close the file if already "
            "open on the system")
        exit()


def adjust_column_widths(ws_object, column_widths):
    # Adjusting Columns Width and Making the Headings Wrap Tezt
    for column_number in range(1, ws_object.max_column + 1):
        column_letter = get_column_letter(column_number)
        column_heading = ws_object[column_letter + "1"].value
        ws_object[column_letter + "1"].alignment = Alignment(wrapText=True)
        if column_heading in column_widths:
            ws_object.column_dimensions[column_letter].width = column_widths[column_heading]
        else:
            ws_object.column_dimensions[column_letter].width = column_widths["Subject"]


# ======================
# MAIN FUNCTIONS TO GET DATA
# ======================

def get_data(result_file, mode):
    # data_format = [Roll No, Gender, Name, Marks, Grade, Pass/Fail, Compartment Subject, Best 5 Subject Marks]
    if mode == "10th":
        return get_data_10th(result_file)
    elif mode == "12th":
        return get_data_12th(result_file)
    else:
        raise ValueError('Invalid format. Valid formats are 10th and 12th.')


def get_data_10th(result_file):
    # Should return the data with headers formatted as
    # [headers,
    #  student-1 details formatted as header,
    #  student-2 details formatted as header,
    #  ...]
    mode = "10th"
    lines = get_lines(result_file)
    lines = filter_lines(lines[:-1], mode)  # Remove the last line containing total candidate numbers
    unique_subject_codes = get_unique_subject_codes(lines[::2])  # Subject Code Lines
    # -- CODE FOR HEADERS --
    headers, unique_subject_names = get_headers(unique_subject_codes, mode)
    # -- CODE FOR STUDENT DETAILS --
    student_details = []
    for index, line in enumerate(lines[::2]):
        individual_student_details = get_individual_student_data(
            list_of_individual_student_lines=lines[index * 2: index * 2 + 2],
            subject_code_order=unique_subject_codes,
            mode=mode
        )
        student_details.append(individual_student_details)

    return headers, student_details, unique_subject_names


# noinspection PyShadowingNames
def get_data_12th(result_file):
    mode = "12th"
    lines = get_lines(result_file)
    lines = filter_lines(lines[:-1], mode)  # Remove the last line containing total candidate numbers
    unique_subject_codes = get_unique_subject_codes(lines[::2])  # Subject Code Lines
    # -- CODE FOR HEADERS --
    headers, unique_subject_names = get_headers(unique_subject_codes, mode)
    # -- CODE FOR STUDENT DETAILS --
    student_details = []
    for index, line in enumerate(lines[::2]):
        individual_student_details = get_individual_student_data(
            list_of_individual_student_lines=lines[index * 2: index * 2 + 2],
            subject_code_order=unique_subject_codes,
            mode=mode
        )
        student_details.append(individual_student_details)

    return headers, student_details, unique_subject_names


def get_individual_student_data(list_of_individual_student_lines, subject_code_order, mode):
    data = []
    first_line_with_compartment = list_of_individual_student_lines[0]

    if mode == '12th':
        first_line_without_compartment = first_line_with_compartment[:144]
    else:
        first_line_without_compartment = first_line_with_compartment[:127]

    second_line = list_of_individual_student_lines[1]

    # Roll No
    data.append(get_roll_no(first_line_without_compartment))
    # Gender
    data.append(get_gender(first_line_without_compartment))
    # Name
    data.append(get_name(first_line_without_compartment))
    # Marks and Grades
    data.extend(get_marks_and_grades(first_line_without_compartment, second_line, subject_code_order))
    # 12th three grades
    if mode == '12th':
        data.extend(get_three_grades(first_line_without_compartment))
    # Result
    data.append(get_result(first_line_with_compartment))
    # Best 5 Percentage
    data.append(best_5_percent(get_marks(second_line)))
    # Compartment Subjects
    data.append(get_compartment_subjects(first_line_with_compartment, mode))
    return data


def exit_program():
    print("The program will exit after 5 seconds.")
    time.sleep(5)
    exit()


if __name__ == '__main__':
    print("Please run main.py and not functions.py")
