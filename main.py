"""
MADE BY PRATHAM GUPTA (https://github.com/PrathamGupta06/cbse-results-analyzer)
"""

import sys
from openpyxl import Workbook, load_workbook
from functions import *
import time
import os

print("-"*64)
print("Welcome to CBSE results analyzer.")
print("-"*64)

# Arguments for the program are -i <input_file> -o <output_file> -c <class>
if len(sys.argv) > 1:
    # Command Line Argument Mode
    # Check if arguments are valid
    if "-i" not in sys.argv or "-o" not in sys.argv or "-c" not in sys.argv:
        print("Invalid Arguments. Please use -i <input_file> -o <output_file> -c <class>")
        exit_program()

    input_file = sys.argv[sys.argv.index("-i") + 1]
    if not validate_input_path(input_file):
        print("ERROR: Invalid Input File Path")
        exit_program()

    output_path_excel = sys.argv[sys.argv.index("-o") + 1]
    if not validate_output_path(output_path_excel):
        print("ERROR: Invalid Output Path")
        exit_program()
    if not output_path_excel.endswith(".xlsx"):   # Add .xlsx if not present
        output_path_excel += ".xlsx"

    mode = sys.argv[sys.argv.index("-c") + 1]
    if mode not in ("10th", "12th"):
        print("ERROR: Invalid Class")
        print("Valid Classes: 10th, 12th")
else:
    # Taking input from user
    input_file = input("\nEnter Input Result File Path (e.g demo/12th.txt): ")
    if not validate_input_path(input_file):
        print("ERROR: Invalid Input File Path")
        exit_program()

    output_path_excel = input("Enter Output Path (e.g 12th.xlsx): ")
    if not validate_output_path(output_path_excel):
        print("ERROR: Invalid Output Path")
        exit_program()
    if not output_path_excel.endswith(".xlsx"):   # Add  .xlsx if not present
        output_path_excel += ".xlsx"

    mode = input("Enter Class (10th or 12th): ")
    if mode not in ("10th", "12th"):
        print("ERROR: Invalid Class")
        print("Valid Classes: 10th, 12th")

start_time = time.time()
# Create Workbook
wb = Workbook()
save_wb(wb, output_path_excel)
wb.close()

column_widths = {
    "Roll No": 20,
    "Gender": 8,
    "Name": 20,
    "Subject": 15,
    "Grade": 6,
    "GR1": 5,
    "GR2": 5,
    "GR3": 5,
    "Result": 8,
    "Best 5 Percentage": 10,
    "Compartment Subject": 15
}

# Get the data from the file
headers, data, AllSubjectNames = get_data(input_file, mode)

# Convert Pandas dataframe to Excel
wb = load_workbook(output_path_excel)
main_sheet = wb["Sheet"]  # Get the main default sheet
main_sheet.title = "Main Result Sheet"  # Rename Sheet
for row in [headers] + data:  # Loop over the data and append it to the main sheet
    main_sheet.append(row)
adjust_column_widths(main_sheet, column_widths)

# Freezing the First Row and First Three Columns of Main Sheet
main_sheet.freeze_panes = "D2"

# Creating Individual Subject Sheets
for SubjectName in AllSubjectNames:
    # Create a new sheet for each subject
    subject_sheet = wb.create_sheet(SubjectName)
    subject_index = headers.index(SubjectName)  # Get the index of the subject in data
    # Append the headers to the subject sheet
    subject_sheet.append([headers[0], headers[1], headers[2], headers[subject_index], headers[subject_index + 1]])
    for row in data:
        # If the student has got marks in the subject
        if row[subject_index] is not None:
            subject_student_row = [row[0], row[1], row[2], row[subject_index], row[subject_index + 1]]
            subject_sheet.append(subject_student_row)
    adjust_column_widths(subject_sheet, column_widths)

# ----------------- Absentees  -----------------
absentee_sheet = wb.create_sheet("Absentees", 1)
absentees = get_absentees(input_file)
absentee_sheet.append(['Roll No', 'Name'])
for absentee in absentees:
    absentee_sheet.append(absentee)
adjust_column_widths(absentee_sheet, column_widths)

# ----------------- Result Analysis  -----------------
# Top 5 Students Analyze
# Splitting the data into Male and Female
male = [student for student in data if student[1] == "M"]  # Add those students whose Gender = "M"
female = [student for student in data if student[1] == "F"]  # Add those students whose Gender = "F"

# Sorting the Male and Female list by Best 5 Percentage
top_n_children = 5  # Number of top children to be shown
best_5_index = headers.index("Best 5 percentage")
top_male = sorted(male, key=lambda student: student[best_5_index], reverse=True)[:top_n_children]
top_female = sorted(female, key=lambda student: student[best_5_index], reverse=True)[:top_n_children]

# Keeping the necessary columns in the list
top_male = [[student[0], student[1], student[2], student[best_5_index]] for student in top_male]
top_female = [[student[0], student[1], student[2], student[best_5_index]] for student in top_female]

# Analyzation related to marks
children_with_full_marks = [["Roll No", "Gender", "Name", "Subject"]]
total_distinctions = 0
all_5_distinctions = 0  # All 5 Subjects Distinction Analyze
all_5_distinction_students = []  # All 5 Subjects Distinction Student List

for student_data in data:
    student_distinctions = 0  # Individual Student Distinctions
    for index, value in enumerate(student_data):
        if type(value) is int or type(value) is float:
            if value >= 75:
                student_distinctions += 1
                total_distinctions += 1
            if value == 100:
                children_with_full_marks.append([student_data[0], student_data[1], student_data[2], headers[index]])
    #  For All 5 Distinctions
    if student_distinctions >= 5:
        all_5_distinctions += 1
        all_5_distinction_students.append([student_data[0], student_data[1], student_data[2]])

# Writing the analysis data to worksheet
wb.create_sheet('Analysis', 2)
analyze_ws = wb['Analysis']

# Children with Full Marks
title = "Children With Full Marks"
append_title(analyze_ws, title, end_column=4, top_row=False)
for i in children_with_full_marks:
    analyze_ws.append(i)


# Male Toppers
title = f"Overall {top_n_children} Male Toppers"
append_title(analyze_ws, title=title, end_column=4)
analyze_ws.append(["Roll No", "Gender", "Name", "Best 5 Percentage"])
for i in top_male:
    analyze_ws.append(i)

# Female Toppers
title = f"Overall {top_n_children} Female Toppers"
append_title(analyze_ws, title=title, end_column=4)
analyze_ws.append(["Roll No", "Gender", "Name", "Best 5 Percentage"])
for i in top_female:
    analyze_ws.append(i)

# Distinctions in all subjects
title = f"No. of Distinctions: {total_distinctions}"
append_title(analyze_ws, title, end_column=4)

# Distinctions in all 5 Subjects
title = f"No. of Distinctions in all 5 subjects: {all_5_distinctions}"
append_title(analyze_ws, title, end_column=4)

# Uncomment the following lines to add the list of students who got all 5 subjects distinctions
'''
append_title(analyze_ws, "Students with distinction in all five subjects", end_column=4)
for i in all_5_distinction_students:
    analyze_ws.append(i)
'''

# Adjust the widths
analyze_ws.column_dimensions["A"].width = column_widths["Roll No"]
analyze_ws.column_dimensions["B"].width = column_widths["Gender"]
analyze_ws.column_dimensions["C"].width = column_widths["Name"]
analyze_ws.column_dimensions["D"].width = column_widths["Subject"]

# Save Workbook
save_wb(wb, output_path_excel)
wb.close()
print("Workbook Saved at", os.path.abspath(output_path_excel))

end_time = time.time()
# Launch Workbook if program is run directly without any arguments
if len(sys.argv) == 1:
    os.startfile(os.path.abspath(output_path_excel))
print(f"Program Completed Successfully and took {round(end_time - start_time, 2)} seconds.")
